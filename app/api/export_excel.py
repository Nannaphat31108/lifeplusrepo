from io import BytesIO
from fastapi import APIRouter,Depends
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill
from app.db.session import get_db
from app.core.security import get_current_user
from app.models.entities import *
router=APIRouter(prefix='/api/export',tags=['Excel Export'])
def finish(wb,name):
 for ws in wb.worksheets:
  ws.freeze_panes='A2';ws.auto_filter.ref=ws.dimensions
  for c in ws[1]:c.font=Font(bold=True,color='FFFFFF');c.fill=PatternFill('solid',fgColor='1F4E78')
  for col in ws.columns:ws.column_dimensions[col[0].column_letter].width=min(max([len(str(x.value or '')) for x in col]+[10])+2,45)
 b=BytesIO();wb.save(b);b.seek(0);return StreamingResponse(b,media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',headers={'Content-Disposition':f'attachment; filename="{name}"'})
def add(wb,name,h,rows):
 ws=wb.create_sheet(name);ws.append(h)
 for r in rows:ws.append(r)
@router.get('/all.xlsx')
def all_data(db:Session=Depends(get_db),u=Depends(get_current_user)):
 wb=Workbook();wb.remove(wb.active)
 add(wb,'Customers',['Code','Name','Contact','Phone','Email'],[[x.customer_code,x.name,x.contact_name or 'ให้ใส่ Data',x.phone or 'ให้ใส่ Data',x.email or 'ให้ใส่ Data'] for x in db.scalars(select(Customer)).all()])
 rows=[]
 for p in db.scalars(select(ProductProject)).all():
  c=db.get(Customer,p.customer_id);ss=db.scalars(select(ProjectSupplementItem).where(ProjectSupplementItem.project_id==p.id)).all() or [None]
  for x in ss:rows.append([p.project_no,c.name if c else 'ให้ใส่ Data',p.product_name,x.supplement_code if x else 'ให้ใส่ Data',x.supplement_name if x else 'ให้ใส่ Data',float(x.amount) if x else '',x.unit if x else '',p.status])
 add(wb,'Projects',['Project','Customer','Product','Supplement Code','Supplement Name','Amount','Unit','Status'],rows)
 add(wb,'Raw Materials',['Code','Trade Name','Ingredient','Halal','FDA'],[[x.material_code,x.trade_name,x.ingredient_name or 'ให้ใส่ Data','Yes' if x.halal else 'No',x.fda_ref or 'ให้ใส่ Data'] for x in db.scalars(select(RawMaterial)).all()])
 add(wb,'Suppliers',['Code','Name','Country','Contact'],[[x.supplier_code,x.name,x.country or 'ให้ใส่ Data',x.contact_name or 'ให้ใส่ Data'] for x in db.scalars(select(Supplier)).all()])
 return finish(wb,'rd_erp_all_data.xlsx')
