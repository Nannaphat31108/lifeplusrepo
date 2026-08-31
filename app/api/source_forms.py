import re
import json, shutil, tempfile, zipfile
from pathlib import Path
from io import BytesIO
from datetime import date, datetime
from fastapi import APIRouter, Depends, HTTPException, Header
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from app.db.session import get_db
from app.core.security import get_current_user
from app.models.entities import SourceFormRecord, Customer, SupplementAlias

router=APIRouter(prefix="/api/source-forms",tags=["Source Forms"])
ROOT=Path(__file__).resolve().parents[2]/"original_forms"

class FormSave(BaseModel):
    record_no:str
    status:str="DRAFT"
    data:dict
    # F-RD-002 only: which month ("YYYY-MM") the user chose to file this
    # record under. F-RD-002.1 only: which person's name the user chose to
    # file it under. Both optional, both user-selected at save time.
    filed_month:str|None=None
    filed_person_name:str|None=None


def require_person_key(
    x_person_key: str | None = Header(default=None, alias="X-Person-Key")
):
    if not x_person_key:
        raise HTTPException(401, "เซสชันหมดอายุ กรุณาล็อกอินใหม่")
    return x_person_key.strip().upper()


def _record_data_dict(x):
    raw=getattr(x,"data_json",None)
    if isinstance(raw,dict):
        return raw
    if isinstance(raw,str):
        try:
            return json.loads(raw)
        except Exception:
            return {}
    raw2=getattr(x,"data",None)
    return raw2 if isinstance(raw2,dict) else {}

def _upsert_customer_from_form(db: Session, data: dict, owner_id: int, record_id: int):
    """Persist the company/customer fields typed into a form into the shared
    Customer master table, instead of leaving them only inside this one
    record's JSON blob. Matches by customer_code first, then by name.
    """
    customer_name = str(data.get("customer_name") or "").strip()
    if not customer_name:
        return
    customer_code = str(data.get("customer_code") or "").strip()
    address = str(data.get("address") or "").strip()
    phone = str(data.get("phone_fax") or data.get("phone") or "").strip()

    customer = None
    if customer_code:
        customer = db.scalar(select(Customer).where(Customer.customer_code == customer_code))
    if not customer:
        customer = db.scalar(select(Customer).where(Customer.name == customer_name))

    if customer:
        customer.name = customer_name
        if address:
            customer.address = address
        if phone:
            customer.phone = phone
        if customer_code and customer.customer_code != customer_code:
            exists = db.scalar(
                select(Customer).where(
                    Customer.customer_code == customer_code,
                    Customer.id != customer.id
                )
            )
            if not exists:
                customer.customer_code = customer_code
    else:
        generated_code = customer_code or f"CUST-{owner_id}-{record_id:06d}"
        db.add(Customer(
            customer_code=generated_code,
            name=customer_name,
            address=address or None,
            phone=phone or None,
        ))


@router.post("/{code}")
def save(
    code: str,
    p: FormSave,
    db: Session = Depends(get_db),
    u=Depends(get_current_user),
    person_key: str = Depends(require_person_key),
):
    if code in {"F-RD-002", "F-RD-002.1"} and not p.data.get("date"):
        p.data["date"] = date.today().isoformat()

    x = SourceFormRecord(
        form_code=code,
        record_no=p.record_no,
        status=p.status,
        payload_json=json.dumps(p.data, ensure_ascii=False, default=str),
        created_by=u.id,
        workspace_user_id=None,
        owner_person_key=person_key,
        filed_month=(p.filed_month.strip() or None) if code == "F-RD-002" and p.filed_month else None,
        filed_person_name=(p.filed_person_name.strip() or None) if code == "F-RD-002.1" and p.filed_person_name else None,
    )
    try:
        db.add(x)
        db.flush()
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Save flush failed: {type(e).__name__}: {e}")

    # Company/customer info typed into a form becomes reusable master data
    # instead of only living inside this record's JSON blob. F-RD-001 has
    # always done this for the customer's name/code; ADMIN-QP (Quotation /
    # Purchase Order) and ADMIN-INVOICE also collect a company name, address
    # and phone, so the same upsert applies to those too.
    if code in {"F-RD-001", "ADMIN-QP", "ADMIN-INVOICE"}:
        _upsert_customer_from_form(db, p.data, owner_id=u.id, record_id=x.id)

    # Store another name/alias against the same extract code.
    for ingredient in (p.data.get("ingredients") or []):
        code_value = str(
            ingredient.get("material_code")
            or ingredient.get("code")
            or ""
        ).strip()
        main_name = str(ingredient.get("name") or "").strip()
        alternate_name = str(ingredient.get("alternate_name") or "").strip()

        if code_value and (main_name or alternate_name):
            alias = db.scalar(
                select(SupplementAlias).where(
                    SupplementAlias.supplement_code == code_value,
                    SupplementAlias.primary_name == (main_name or alternate_name)
                )
            )
            if not alias:
                db.add(SupplementAlias(
                    supplement_code=code_value,
                    primary_name=main_name or alternate_name,
                    alternate_name=alternate_name or None
                ))
            elif alternate_name:
                alias.alternate_name = alternate_name

    try:
        db.commit()
        db.refresh(x)
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Save failed: {type(e).__name__}: {e}")
    return {"id": x.id, "record_no": x.record_no, "owner": u.full_name}

@router.get("/formula-link/{formula_no}")
def formula_link_for_qp(
    formula_no: str,
    db: Session = Depends(get_db),
    u=Depends(get_current_user),
    person_key: str = Depends(require_person_key),
):
    """Return only the fields allowed to flow from F-RD-002 into ADMIN-QP.

    QP intentionally receives only ingredient name, quantity (mg), and origin country.
    Supplier, material code, price, halal, FDA and other R&D-only fields are excluded.
    """
    wanted = re.sub(r"\s+", "", str(formula_no or "")).upper()
    # VLOOKUP-like lookup for ADMIN-QP: formula numbers are shared internal
    # references, so ADMIN must be able to resolve a formula created by R&D.
    # Only the slim quotation-safe fields below are returned.
    rows = db.scalars(
        select(SourceFormRecord)
        .where(SourceFormRecord.form_code == "F-RD-002")
        .order_by(SourceFormRecord.id.desc())
    ).all()

    for rec in rows:
        try:
            data = json.loads(rec.payload_json or "{}")
        except Exception:
            continue
        current = re.sub(r"\s+", "", str(data.get("formula_no") or "")).upper()
        if current != wanted:
            continue

        def slim(items, limit):
            out=[]
            for x in (items or [])[:limit]:
                if not isinstance(x, dict):
                    continue
                out.append({
                    "name": x.get("name") or x.get("ingredient_name") or "",
                    "quantity_mg": x.get("quantity_mg") if x.get("quantity_mg") not in (None, "") else x.get("quantity"),
                    "origin": x.get("import_country") or x.get("origin") or "",
                })
            return out

        return {
            "formula_no": data.get("formula_no") or formula_no,
            "customer_name": data.get("customer_name") or "",
            "product_name": data.get("product_name") or data.get("formula_name") or "",
            "ingredients": slim(data.get("ingredients"), 9),
            "inactive_ingredients": slim(data.get("inactive_ingredients"), 7),
        }

    raise HTTPException(404, "ไม่พบรหัสสูตรนี้ในไฟล์สูตร F-RD-002")

@router.get("/{code}")
def list_records(
    code: str,
    month: str | None = None,
    person_name: str | None = None,
    db: Session = Depends(get_db),
    u=Depends(get_current_user),
    person_key: str = Depends(require_person_key),
):
    """List this person's own records for a form.

    F-RD-002 records can optionally be filtered by `month` (the "YYYY-MM"
    they were filed under); F-RD-002.1 records can optionally be filtered by
    `person_name` (the name they were filed under) -- both independent of
    who is actually logged in.
    """
    conditions = [
        SourceFormRecord.form_code == code,
        SourceFormRecord.created_by == u.id,
        SourceFormRecord.owner_person_key == person_key,
    ]
    if code == "F-RD-002" and month:
        conditions.append(SourceFormRecord.filed_month == month)
    if code == "F-RD-002.1" and person_name:
        conditions.append(SourceFormRecord.filed_person_name == person_name)

    rows = db.scalars(
        select(SourceFormRecord)
        .where(*conditions)
        .order_by(SourceFormRecord.id.desc())
    ).all()
    return [{
        "id": x.id,
        "record_no": x.record_no,
        "status": x.status,
        "data": json.loads(x.payload_json or "{}"),
        "created_at": x.created_at,
        "owner": u.full_name,
        "filed_month": x.filed_month,
        "filed_person_name": x.filed_person_name,
    } for x in rows]

@router.get("/record/{record_id}")
def get_record(
    record_id: int,
    db: Session = Depends(get_db),
    u=Depends(get_current_user),
    person_key: str = Depends(require_person_key),
):
    x = db.get(SourceFormRecord, record_id)
    if not x or x.created_by != u.id or x.owner_person_key != person_key:
        raise HTTPException(404, "Record not found")
    return {
        "id": x.id,
        "form_code": x.form_code,
        "record_no": x.record_no,
        "status": x.status,
        "data": json.loads(x.payload_json or "{}"),
        "filed_month": x.filed_month,
        "filed_person_name": x.filed_person_name,
    }


@router.put("/record/{record_id}")
def update_record(
    record_id: int,
    p: FormSave,
    db: Session = Depends(get_db),
    u=Depends(get_current_user),
    person_key: str = Depends(require_person_key),
):
    x = db.get(SourceFormRecord, record_id)
    if not x or x.created_by != u.id or x.owner_person_key != person_key:
        raise HTTPException(404, "Record not found")

    if x.form_code in {"F-RD-002", "F-RD-002.1"} and not p.data.get("date"):
        p.data["date"] = date.today().isoformat()

    x.record_no = p.record_no
    x.status = p.status
    x.owner_person_key = person_key
    x.payload_json = json.dumps(p.data, ensure_ascii=False, default=str)
    if x.form_code == "F-RD-002" and p.filed_month is not None:
        x.filed_month = p.filed_month.strip() or None
    if x.form_code == "F-RD-002.1" and p.filed_person_name is not None:
        x.filed_person_name = p.filed_person_name.strip() or None

    # F-RD-001 always updates/syncs Customer master data.
    if x.form_code == "F-RD-001":
        customer_name = str(p.data.get("customer_name") or "").strip()
        customer_code = str(p.data.get("customer_code") or "").strip()
        if customer_name:
            customer = None
            if customer_code:
                customer = db.scalar(select(Customer).where(Customer.customer_code == customer_code))
            if not customer:
                customer = db.scalar(select(Customer).where(Customer.name == customer_name))
            if customer:
                customer.name = customer_name
            else:
                db.add(Customer(
                    customer_code=customer_code or f"CUST-{u.id}-{x.id:06d}",
                    name=customer_name
                ))

    # Same extract code can keep primary name + another name.
    for ing in (p.data.get("ingredients") or []):
        sc = str(ing.get("material_code") or ing.get("code") or "").strip()
        pn = str(ing.get("name") or "").strip()
        an = str(ing.get("alternate_name") or "").strip()
        if sc and (pn or an):
            alias = db.scalar(
                select(SupplementAlias).where(
                    SupplementAlias.supplement_code == sc,
                    SupplementAlias.primary_name == (pn or an)
                )
            )
            if not alias:
                db.add(SupplementAlias(
                    supplement_code=sc,
                    primary_name=pn or an,
                    alternate_name=an or None
                ))
            elif an:
                alias.alternate_name = an

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Update failed: {type(e).__name__}: {e}")
    return {"id": x.id, "record_no": x.record_no, "owner": u.full_name, "updated": True}


def put(ws,cell,value):
    if value is None or value == "":
        return

    target=ws[cell]

    # If address points to a child cell inside a merged range,
    # write to the top-left owner cell instead.
    if isinstance(target, MergedCell):
        for rng in ws.merged_cells.ranges:
            if cell in rng:
                target=ws.cell(rng.min_row, rng.min_col)
                break

    target.value=value

def _shift_formula_addr(addr:str, production:bool, ingredient_count:int):
    capacity=12 if production else 20
    if ingredient_count<=capacity:
        return addr
    extra=ingredient_count-capacity
    m=re.match(r"^([A-Z]+)(\d+)$",addr or "")
    if not m:
        return addr
    col,row=m.group(1),int(m.group(2))
    insert_at=28 if production else 36
    if row>=insert_at:
        row+=extra
    return f"{col}{row}"

def fill_tester_qty_formula(ws,d):
    """F-RD-002.1 AP31 ("จำนวน Tester") was a hardcoded master formula
    "=30/0.981". The "/0.981" did not generalize to a custom count (confirmed
    against a real record: 100 testers on a 10mg/200mg formula must give
    exactly 1000mg/20000mg, not 1019.36/20387.3 — i.e. pack_mg is simply
    quantity_mg * tester count), so this now writes the plain count itself.
    AO16..AO27/AO28 keep referencing AP31 by cell, so they still compute
    correctly off whatever count is entered.
    """
    ingredient_count=int(d.get("ingredient_count") or len(d.get("ingredients",[]) or []) or 0)
    raw=(d.get("manual_cells") or {}).get("AP31")
    try:
        qty=float(raw) if raw not in (None,"") else 30
        if qty<=0:
            qty=30
    except (TypeError,ValueError):
        qty=30
    target=_shift_formula_addr("AP31",True,ingredient_count)
    put(ws,target,qty)


def fill_margin_percent_formula(ws,d):
    """New field: overall margin % = K35*100/K34 (profit including
    packaging cost / selling price), written to Z36 — distinct from the
    master's existing Z35 which excludes packaging cost (AE31).
    """
    ingredient_count=int(d.get("ingredient_count") or len(d.get("ingredients",[]) or []) or 0)
    target=_shift_formula_addr("Z36",True,ingredient_count)
    k34=_shift_formula_addr("K34",True,ingredient_count)
    k35=_shift_formula_addr("K35",True,ingredient_count)
    put(ws,target,f"={k35}*100/{k34}")


def fill_manual_cells(ws,d,production=False):
    ingredient_count=int(d.get("ingredient_count") or len(d.get("ingredients",[]) or []) or 0)
    for cell,value in (d.get("manual_cells") or {}).items():
        try:
            target=_shift_formula_addr(cell,production,ingredient_count)
            put(ws,target,value)
        except Exception:
            pass

def fill_001(ws,d):
    put(ws,"H4",d.get("customer_name")); put(ws,"H6",d.get("customer_code"))
    put(ws,"H9",d.get("product_category")); put(ws,"H17",d.get("objective")); put(ws,"H23",d.get("product_detail"))
    items=d.get("ingredients",[])
    for i,x in enumerate(items[:10]):
        row=35+(i if i<5 else i-5); namecell=("C" if i<5 else "Z")+str(row); amtcell=("R" if i<5 else "AO")+str(row)
        put(ws,namecell,x.get("name"));put(ws,amtcell,x.get("amount"))
    put(ws,"N41",d.get("order_capsule"));put(ws,"R41",d.get("order_sachet"));put(ws,"U41",d.get("order_tablet"))
    for i,x in enumerate(d.get("formula_rates",[])[:5],42):
        put(ws,f"AA{i}",x.get("formula_no"));put(ws,f"AL{i}",x.get("price"))

from copy import copy
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange

def _copy_excel_row_style(ws, source_row:int, target_row:int):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

    for col in range(1, ws.max_column + 1):
        src=ws.cell(source_row,col)
        dst=ws.cell(target_row,col)

        # MergedCell children are read-only; skip them entirely.
        if isinstance(src, MergedCell) or isinstance(dst, MergedCell):
            continue

        if src.has_style:
            dst._style=copy(src._style)

        dst.font=copy(src.font)
        dst.fill=copy(src.fill)
        dst.border=copy(src.border)
        dst.alignment=copy(src.alignment)
        dst.protection=copy(src.protection)
        dst.number_format=src.number_format

def expand_formula_inactive_rows(ws, active_extra:int, inactive_extra:int):
    """Grow F-RD-002's Inactive Ingredient section past its 3-row template
    (rows 39-41, before any active-ingredient expansion). Mirrors
    expand_formula_ingredient_rows, but for the inactive block: it inserts
    rows right after the last inactive template row and shifts the
    subtotal/total/summary/sign rows below it down accordingly, so an
    "unlimited inactive ingredients" count (like the existing unlimited
    active-ingredient system) survives the real Excel export intact instead
    of overwriting the subtotal rows that follow.
    """
    if inactive_extra<=0:
        return

    insert_at=42+active_extra
    template_row=41+active_extra

    original_merges=[CellRange(str(rng)) for rng in list(ws.merged_cells.ranges)]

    template_merges=[]
    for cr in original_merges:
        if cr.min_row==template_row and cr.max_row==template_row:
            template_merges.append(CellRange(str(cr)))

    for cr in original_merges:
        if cr.min_row>=insert_at:
            try:
                ws.unmerge_cells(str(cr))
            except Exception:
                pass

    ws.insert_rows(insert_at, amount=inactive_extra)

    for offset in range(inactive_extra):
        row=insert_at+offset
        _copy_excel_row_style(ws,template_row,row)

        for tcr in template_merges:
            new_cr=CellRange(
                min_col=tcr.min_col,
                min_row=row,
                max_col=tcr.max_col,
                max_row=row
            )
            try:
                ws.merge_cells(str(new_cr))
            except Exception:
                pass

    for cr in original_merges:
        if cr.min_row>=insert_at:
            shifted=CellRange(str(cr))
            shifted.shift(row_shift=inactive_extra,col_shift=0)
            try:
                ws.merge_cells(str(shifted))
            except Exception:
                pass

def expand_formula_ingredient_rows(ws, production:bool, ingredient_count:int):
    capacity=12 if production else 20
    if ingredient_count<=capacity:
        return

    extra=ingredient_count-capacity
    insert_at=28 if production else 36
    template_row=27 if production else 35

    # Capture all merge ranges before changing the worksheet.
    original_merges=[CellRange(str(rng)) for rng in list(ws.merged_cells.ranges)]

    # Capture one-row merge pattern of the final ingredient template row.
    template_merges=[]
    for cr in original_merges:
        if cr.min_row==template_row and cr.max_row==template_row:
            template_merges.append(CellRange(str(cr)))

    # Unmerge ranges at/below the insert position before inserting rows.
    for cr in original_merges:
        if cr.min_row>=insert_at:
            try:
                ws.unmerge_cells(str(cr))
            except Exception:
                pass

    ws.insert_rows(insert_at, amount=extra)

    # Copy visual style from last original ingredient row.
    for offset in range(extra):
        row=insert_at+offset
        _copy_excel_row_style(ws,template_row,row)

        # Number column: safe top-left write only.
        put(ws,f"B{row}",capacity+offset+1)

        # Recreate same-row merges for each inserted ingredient row.
        for tcr in template_merges:
            new_cr=CellRange(
                min_col=tcr.min_col,
                min_row=row,
                max_col=tcr.max_col,
                max_row=row
            )
            try:
                ws.merge_cells(str(new_cr))
            except Exception:
                pass

    # Restore all original merges below inserted section, shifted down.
    for cr in original_merges:
        if cr.min_row>=insert_at:
            shifted=CellRange(str(cr))
            shifted.shift(row_shift=extra,col_shift=0)
            try:
                ws.merge_cells(str(shifted))
            except Exception:
                pass

def fill_formula(ws,d,production=False):
    put(ws,"I6" if production else "I5",d.get("customer_name"))
    put(ws,"AJ6" if production else "AJ5",d.get("formula_no"))
    put(ws,"I8" if production else "I7",d.get("product_type"))
    put(ws,"AJ8" if production else "AJ7",d.get("date"))
    put(ws,"I10" if production else "I9",d.get("product_name_fda"))
    put(ws,"AJ10" if production else "AJ9",d.get("salesperson"))
    put(ws,"I12" if production else "I11",d.get("order_quantity"))
    put(ws,"Q12" if production else "T11",d.get("order_unit"))

    ingredients=d.get("ingredients",[]) or []
    inactive=d.get("inactive_ingredients",[]) or []
    count=max(int(d.get("ingredient_count") or 0),len(ingredients))

    expand_formula_ingredient_rows(ws,production,count)

    if production:
        # Actual F-RD-002.1 ingredient template is rows 16-27 (12 rows).
        capacity=12
        extra=max(0,count-capacity)
        total_row=28+extra
        packaging_row=31+extra
        cost_row=33+extra
        sale_row=34+extra
        profit_row=35+extra
        tester_cost_row=36+extra
        sign_row=41+extra

        last_active=15+max(count,1)

        for idx,x in enumerate(ingredients):
            row=16+idx
            put(ws,f"D{row}",x.get("name"))
            put(ws,f"P{row}",x.get("quantity_mg"))
            put(ws,f"AA{row}",x.get("price_kg"))
            put(ws,f"AI{row}",x.get("supplier"))
            put(ws,f"AM{row}",x.get("material_code"))
            put(ws,f"AN{row}",x.get("price_pack"))
            put(ws,f"AO{row}",x.get("pack_mg"))
            put(ws,f"AP{row}",x.get("quantity_g"))

            # Calculation Master ingredient rules.
            put(ws,f"V{row}",f"=SUM(P{row}*$I$12/1000000)")
            put(ws,f"Z{row}",f"=P{row}*100/$P${total_row}")
            put(ws,f"AE{row}",f"=SUM(AA{row}/1000000*P{row})")

            # Keep formula-production packaging calculations.
            put(ws,f"AO{row}",f"=P{row}*AP{packaging_row}")
            put(ws,f"AP{row}",f"=AO{row}/1000")
            put(ws,f"AQ{row}",f"=SUM(AA{row}/1000000*AO{row})")

        put(ws,f"P{total_row}",f"=SUM(P16:U{last_active})")
        put(ws,f"V{total_row}",f"=SUM(V16:Y{last_active})")
        put(ws,f"Z{total_row}",f"=SUM(Z16:Z{last_active})")

        # Existing formula-production summary structure.
        put(ws,f"K{cost_row}",f"=SUM(AE16:AH{packaging_row})")
        put(ws,f"AO{cost_row}",f"=SUM(I12*K{cost_row})")
        put(ws,f"AO{sale_row}",f"=SUM(I12*K{sale_row})")
        put(ws,f"K{profit_row}",f"=SUM(K{sale_row}-K{cost_row})")
        put(ws,f"AO{profit_row}",f"=SUM(AO{sale_row}-AO{cost_row})")
        put(ws,f"AI{sign_row}",f"=AJ8")

        put(ws,f"B{38+extra}",d.get("rate_text"))
        put(ws,f"B{47+extra}",d.get("formula_note"))
        put(ws,f"AI{40+extra}",d.get("signature_name"))

    else:
        # F-RD-002 current layout: 20 active rows, then fixed 3 inactive rows.
        capacity=20
        extra=max(0,count-capacity)

        # Inactive Ingredient section: same "unlimited count" treatment as the
        # active section above (template capacity 3 rows). Insert real rows
        # into the workbook for anything beyond that, then shift every row
        # constant below it accordingly.
        inactive_extra=max(0,len(inactive)-3)
        if inactive_extra:
            expand_formula_inactive_rows(ws,extra,inactive_extra)

        active_subtotal_row=36+extra
        inactive_header_row=37+extra
        inactive_label_row=38+extra
        inactive_start=39+extra
        inactive_end=inactive_start+max(len(inactive),3)-1
        inactive_subtotal_row=42+extra+inactive_extra
        total_row=43+extra+inactive_extra

        qty_summary_row=44+extra+inactive_extra
        prod_summary_row=45+extra+inactive_extra
        cost_row=47+extra+inactive_extra
        sale_row=48+extra+inactive_extra
        profit_row=49+extra+inactive_extra
        date_sign_row=54+extra+inactive_extra

        active_last=15+max(count,1)

        for idx,x in enumerate(ingredients):
            row=16+idx
            put(ws,f"D{row}",x.get("name"))
            put(ws,f"T{row}",x.get("quantity_mg"))
            put(ws,f"AE{row}",x.get("price_kg"))
            put(ws,f"AM{row}",x.get("supplier"))
            put(ws,f"AR{row}",x.get("import_country"))
            put(ws,f"AS{row}",x.get("material_code"))
            put(ws,f"AT{row}",x.get("halal"))

            # Exact master formulas.
            put(ws,f"Z{row}",f"=SUM(T{row}*$I$11/1000000)")
            put(ws,f"AD{row}",f"=T{row}*100/$T${total_row}")
            put(ws,f"AI{row}",f"=SUM(AE{row}/1000000*T{row})")

        # Active subtotal.
        put(ws,f"T{active_subtotal_row}",f"=SUM(T16:Y{active_last})")
        put(ws,f"Z{active_subtotal_row}",f"=SUM(Z16:AC{active_last})")
        put(ws,f"AD{active_subtotal_row}",f"=SUM(AD16:AD{active_last})")

        # Inactive ingredient rows.
        for j,x in enumerate(inactive):
            row=inactive_start+j
            put(ws,f"B{row}",j+1)
            put(ws,f"D{row}",x.get("name"))
            put(ws,f"T{row}",x.get("quantity_mg"))
            put(ws,f"AE{row}",x.get("price_kg"))
            put(ws,f"AM{row}",x.get("supplier"))
            put(ws,f"AR{row}",x.get("import_country"))
            put(ws,f"AS{row}",x.get("material_code"))
            put(ws,f"AT{row}",x.get("halal"))

            put(ws,f"Z{row}",f"=SUM(T{row}*$I$11/1000000)")
            put(ws,f"AD{row}",f"=T{row}*100/$T${total_row}")
            put(ws,f"AI{row}",f"=SUM(AE{row}/1000000*T{row})")

        actual_inactive_last=inactive_start+max(len(inactive),1)-1

        # Same subtotal/total hierarchy as Calculation Master.
        put(ws,f"T{inactive_subtotal_row}",f"=SUM(T{inactive_start}:Y{actual_inactive_last})")
        put(ws,f"Z{inactive_subtotal_row}",f"=SUM(Z{inactive_start}:AC{actual_inactive_last})")
        put(ws,f"AD{inactive_subtotal_row}",f"=SUM(AD{inactive_start}:AD{actual_inactive_last})")

        put(ws,f"T{total_row}",f"=T{active_subtotal_row}+T{inactive_subtotal_row}")
        put(ws,f"Z{total_row}",f"=SUM(Z{active_subtotal_row}+Z{inactive_subtotal_row})")
        put(ws,f"AD{total_row}",f"=SUM(AD{active_subtotal_row}+AD{inactive_subtotal_row})")

        put(ws,f"K{qty_summary_row}",f"=SUM(T{total_row})")
        put(ws,f"K{prod_summary_row}",f"=SUM(Z{total_row})")

        # Exact master costing rule includes *120.
        put(ws,f"K{cost_row}",f"=SUM(AI16:AL{actual_inactive_last})*120")
        put(ws,f"AO{cost_row}",f"=SUM(I11*K{cost_row})")
        put(ws,f"AO{sale_row}",f"=SUM(I11*K{sale_row})")
        put(ws,f"K{profit_row}",f"=SUM(K{sale_row}-K{cost_row})")
        put(ws,f"AO{profit_row}",f"=SUM(AO{sale_row}-AO{cost_row})")
        put(ws,f"AJ{date_sign_row}",f"=AG7")

        put(ws,f"B{51+extra+inactive_extra}",d.get("rate_text"))
        put(ws,f"S{51+extra+inactive_extra}",d.get("formula_note"))
        put(ws,f"AJ{53+extra+inactive_extra}",d.get("signature_name"))

def fill_003(ws,d):
    put(ws,"C5",d.get("quotation_no"));put(ws,"H5",d.get("formula_no"));put(ws,"C7",d.get("customer_name"));put(ws,"H7",d.get("receipt_no"))
    put(ws,"C9",d.get("customer_needed"));put(ws,"C13",d.get("characteristic"));put(ws,"C15",d.get("packaging"))
    put(ws,"C16",d.get("quantity"));put(ws,"H16",d.get("delivery_date"));put(ws,"C18",d.get("tester_type"));put(ws,"E18",d.get("price"));put(ws,"H18",d.get("payin_ref"));put(ws,"A20",d.get("requester"));put(ws,"F20",d.get("rd_maker"))
def fill_004(ws,d):
    put(ws,"C4",d.get("customer_name"));put(ws,"H4",d.get("customer_code"));put(ws,"C6",d.get("op_no"));put(ws,"H6",d.get("formula_no"))
    put(ws,"C8",d.get("formula_name"));put(ws,"C9",d.get("product_name"))
    for i,x in enumerate(d.get("rates",[])[:10],12):
        put(ws,f"A{i}",x.get("quantity"));put(ws,f"E{i}",x.get("price_unit"));put(ws,f"H{i}",x.get("note"))


def fill_admin_qp(ws,d):
    """Fill the real ADMIN-QP master workbook without changing its layout.

    The original workbook remains the visual/calculation master. Linked R&D data is
    deliberately limited to ingredient name, quantity mg, and origin country.
    """
    # Header fields in the original QP sheet.
    put(ws,"I8",d.get("customer_name"))
    put(ws,"AP8",d.get("quotation_no") or d.get("ref_no"))
    put(ws,"AP10",d.get("date"))
    put(ws,"I10",d.get("address"))
    put(ws,"I12",d.get("phone_fax"))
    put(ws,"I14",d.get("product_name"))
    put(ws,"AP17",d.get("formula_no"))
    # Dates are editable only inside the parentheses below the English titles.
    # Keep D60 and every other master element unchanged.
    sales_sign_date=str(d.get("sales_signature_date") or "").strip()
    admin_sign_date=str(d.get("admin_signature_date") or "").strip()
    if sales_sign_date:
        put(ws,"W61",sales_sign_date)
    if admin_sign_date:
        put(ws,"AM61",admin_sign_date)

    # The exact master has 9 Active slots and 7 Inactive slots.
    ingredients=d.get("qp_ingredients") or []
    active=ingredients[:9]
    inactive=ingredients[9:16]

    def ingredient_text(x):
        name=str((x or {}).get("ingredient_name") or (x or {}).get("name") or "").strip()
        origin=str((x or {}).get("origin") or (x or {}).get("import_country") or "").strip()
        if name and origin:
            return f"{name} ประเทศ {origin}"
        return name or (f"ประเทศ {origin}" if origin else "")

    for offset,x in enumerate(active):
        row=20+offset
        put(ws,f"D{row}",ingredient_text(x))
        put(ws,f"T{row}",x.get("quantity_mg"))

    for offset,x in enumerate(inactive):
        row=22+offset
        put(ws,f"AB{row}",ingredient_text(x))
        put(ws,f"AR{row}",x.get("quantity_mg"))

    # Quotation/Purchase Order lines use the original cells and formulas.
    for offset,x in enumerate((d.get("qp_lines") or [])[:13]):
        row=33+offset
        put(ws,f"D{row}",x.get("job_code"))
        put(ws,f"I{row}",x.get("description"))
        put(ws,f"V{row}",x.get("pack_qty"))
        put(ws,f"X{row}",x.get("pack_unit_text"))
        put(ws,f"AE{row}",x.get("quantity"))
        put(ws,f"AJ{row}",x.get("unit"))
        put(ws,f"AL{row}",x.get("unit_price"))
        if x.get("amount") not in (None, ""):
            put(ws,f"AR{row}",x.get("amount"))

    # Optional manual discount; all remaining totals stay as formulas from MASTER.
    if d.get("discount") not in (None, ""):
        put(ws,"AR48",d.get("discount"))

    # หมายเหตุ (notes/remarks) -- the master template has no dedicated cell for
    # this, so it goes in the one genuinely free row between the amount-in-words
    # line (B53:AK54) and the signature block (row 56). Note: the ADMIN-QP/
    # ADMIN-INVOICE export path (_admin_qp_export_preserve_master) only diffs
    # and patches cell *values* to protect the master's embedded images/styling,
    # so a style change here (e.g. wrap_text) would not survive into the
    # exported file -- the value itself does, and simply overflows visually
    # into the empty cells to its right like any other long unmerged text.
    notes=str(d.get("notes") or "").strip()
    if notes:
        put(ws,"B55",notes)

    # Signature/signatory fields intentionally omitted from ADMIN-QP.


def _write_formula_fda_column(ws,d,production=False):
    """Write FDA No. in the far-right FDA column using the real formula rows."""
    from openpyxl.utils import get_column_letter

    col=get_column_letter(ws.max_column+1)
    put(ws,f"{col}15","FDA No.")

    active=d.get("ingredients") or []
    for row,x in enumerate(active,16):
        put(ws,f"{col}{row}",x.get("fda_no") or "")

    inactive=d.get("inactive_ingredients") or []
    inactive_start=34 if production else 39
    for offset,x in enumerate(inactive):
        put(ws,f"{col}{inactive_start+offset}",x.get("fda_no") or "")

def _normalize_material_code(value):
    import re
    raw=str(value or "").strip().upper()

    m=re.match(r"^([A-Z]+)0*(\d{1,4})(?:[^\d].*)?$",raw)
    if m:
        return m.group(1)+m.group(2)[-4:].zfill(4)

    m=re.search(r"([A-Z]+)0*(\d{1,4})",raw)
    if m:
        return m.group(1)+m.group(2)[-4:].zfill(4)

    return re.sub(r"\s+","",raw)


def _normalize_formula_material_codes(data):
    for key in ("ingredients","inactive_ingredients"):
        for item in data.get(key) or []:
            if not isinstance(item,dict):
                continue

            variant=str(
                item.get("variant_code") or item.get("material_code") or ""
            ).strip().upper()

            item["variant_code"]=variant
            item["material_code"]=_normalize_material_code(
                item.get("material_code") or variant
            )

    return data


def _xlsx_sheet_part_for_name(xlsx_path: Path, sheet_name: str) -> str:
    """Return the worksheet XML part for a visible sheet name."""
    import xml.etree.ElementTree as ET

    NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    NS_PKG = "http://schemas.openxmlformats.org/package/2006/relationships"

    with zipfile.ZipFile(xlsx_path, "r") as zf:
        wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
        rel_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))

        rel_map = {
            rel.attrib.get("Id"): rel.attrib.get("Target")
            for rel in rel_root.findall(f"{{{NS_PKG}}}Relationship")
        }

        for sh in wb_root.findall(f".//{{{NS_MAIN}}}sheet"):
            if sh.attrib.get("name") != sheet_name:
                continue
            rid = sh.attrib.get(f"{{{NS_REL}}}id")
            target = rel_map.get(rid)
            if not target:
                break
            if target.startswith("/"):
                return target.lstrip("/")
            return "xl/" + target.lstrip("/")

    raise ValueError(f"Sheet not found in xlsx package: {sheet_name}")


def _patch_xlsx_cells_preserve_package(master_path: Path, sheet_name: str, updates: dict) -> bytes:
    """Patch cell values directly in the XLSX ZIP package.

    This intentionally avoids saving the visual master through openpyxl.  The
    original drawings, logos, media files, relationships, print settings, and
    every other package part therefore remain byte-for-byte untouched except
    for the target worksheet XML.
    """
    import math
    import xml.etree.ElementTree as ET

    NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ET.register_namespace("", NS)

    sheet_part = _xlsx_sheet_part_for_name(master_path, sheet_name)

    with zipfile.ZipFile(master_path, "r") as zin:
        original_sheet = zin.read(sheet_part)
        root = ET.fromstring(original_sheet)
        sheet_data = root.find(f"{{{NS}}}sheetData")
        if sheet_data is None:
            raise ValueError("worksheet has no sheetData")

        rows = {}
        cells = {}
        for row in sheet_data.findall(f"{{{NS}}}row"):
            rnum = int(row.attrib.get("r", "0") or 0)
            rows[rnum] = row
            for cell in row.findall(f"{{{NS}}}c"):
                ref = cell.attrib.get("r")
                if ref:
                    cells[ref] = cell

        def split_ref(ref):
            m = re.match(r"^([A-Z]+)(\d+)$", str(ref or "").upper())
            if not m:
                raise ValueError(f"invalid cell reference: {ref}")
            return m.group(1), int(m.group(2))

        def col_num(col):
            n = 0
            for ch in col:
                n = n * 26 + (ord(ch) - 64)
            return n

        def ensure_row(rnum):
            row = rows.get(rnum)
            if row is not None:
                return row
            row = ET.Element(f"{{{NS}}}row", {"r": str(rnum)})
            inserted = False
            for idx, existing in enumerate(list(sheet_data)):
                er = int(existing.attrib.get("r", "0") or 0)
                if er > rnum:
                    sheet_data.insert(idx, row)
                    inserted = True
                    break
            if not inserted:
                sheet_data.append(row)
            rows[rnum] = row
            return row

        def ensure_cell(ref):
            cell = cells.get(ref)
            if cell is not None:
                return cell
            col, rnum = split_ref(ref)
            row = ensure_row(rnum)
            cell = ET.Element(f"{{{NS}}}c", {"r": ref})
            wanted = col_num(col)
            inserted = False
            for idx, existing in enumerate(list(row)):
                eref = existing.attrib.get("r")
                if not eref:
                    continue
                ecol, _ = split_ref(eref)
                if col_num(ecol) > wanted:
                    row.insert(idx, cell)
                    inserted = True
                    break
            if not inserted:
                row.append(cell)
            cells[ref] = cell
            return cell

        def clear_value_nodes(cell):
            for tag in ("f", "v", "is"):
                node = cell.find(f"{{{NS}}}{tag}")
                if node is not None:
                    cell.remove(node)

        for ref, value in updates.items():
            if value is None or value == "":
                continue
            ref = str(ref).upper()
            cell = ensure_cell(ref)
            clear_value_nodes(cell)

            if isinstance(value, bool):
                cell.attrib["t"] = "b"
                v = ET.SubElement(cell, f"{{{NS}}}v")
                v.text = "1" if value else "0"
            elif isinstance(value, (int, float)) and not isinstance(value, bool):
                if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
                    value = str(value)
                else:
                    cell.attrib.pop("t", None)
                    v = ET.SubElement(cell, f"{{{NS}}}v")
                    v.text = str(value)
                    continue
                cell.attrib["t"] = "inlineStr"
                is_node = ET.SubElement(cell, f"{{{NS}}}is")
                t = ET.SubElement(is_node, f"{{{NS}}}t")
                t.text = value
            else:
                text = str(value)
                if text.startswith("="):
                    cell.attrib.pop("t", None)
                    f = ET.SubElement(cell, f"{{{NS}}}f")
                    f.text = text[1:]
                else:
                    cell.attrib["t"] = "inlineStr"
                    is_node = ET.SubElement(cell, f"{{{NS}}}is")
                    t = ET.SubElement(is_node, f"{{{NS}}}t")
                    if text[:1].isspace() or text[-1:].isspace():
                        t.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
                    t.text = text

        patched = ET.tostring(root, encoding="utf-8", xml_declaration=True)

        out = BytesIO()
        with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for info in zin.infolist():
                data = patched if info.filename == sheet_part else zin.read(info.filename)
                zout.writestr(info, data)
        return out.getvalue()


def _admin_qp_export_preserve_master(master_path: Path, data: dict, is_invoice: bool = False) -> bytes:
    """Build ADMIN-QP (or the cloned ADMIN-INVOICE) from the master while
    preserving every embedded image. Both forms share the exact same layout
    and cells — the invoice variant only overwrites the title text.
    """
    wb = load_workbook(master_path, data_only=False)
    if "ใบราคา" not in wb.sheetnames:
        raise ValueError("Sheet not found: ใบราคา")
    ws = wb["ใบราคา"]

    before = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            before[cell.coordinate] = cell.value

    fill_admin_qp(ws, data)
    if is_invoice:
        # Same document, relabeled: "ใบเสนอราคา / ใบสั่งซื้อ" -> "ใบแจ้งหนี้".
        put(ws, "B5", "ใบแจ้งหนี้")
        put(ws, "B7", "Invoice")
    fill_manual_cells(ws, data, False)

    updates = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            old = before.get(cell.coordinate)
            if cell.value != old:
                updates[cell.coordinate] = cell.value

    return _patch_xlsx_cells_preserve_package(master_path, "ใบราคา", updates)

def _verify_master_media_preserved(master_path: Path, output_bytes: bytes):
    """Verify embedded media survives export without relying on ZIP member names.

    openpyxl legitimately renames/re-encodes workbook images when saving, e.g.
    ``xl/media/image.png`` can become ``xl/media/image1.png`` and JPG files can
    become JPEG files.  Comparing exact member names therefore creates false
    failures even when every logo/image is still embedded.
    """
    try:
        def _media_entries(xlsx_zip):
            entries=[]
            for name in xlsx_zip.namelist():
                if not name.startswith("xl/media/") or name.endswith("/"):
                    continue
                data=xlsx_zip.read(name)
                entries.append((name, len(data)))
            return entries

        with zipfile.ZipFile(master_path, "r") as src_zip:
            master_media=_media_entries(src_zip)

        if not master_media:
            return

        with zipfile.ZipFile(BytesIO(output_bytes), "r") as out_zip:
            output_media=_media_entries(out_zip)

        # The important invariant is that all embedded media objects survive
        # and none are empty. Exact filenames/extensions are not stable across
        # an openpyxl load/save cycle.
        empty=[name for name,size in output_media if size <= 0]
        if len(output_media) < len(master_media) or empty:
            details=[f"expected_media>={len(master_media)}", f"actual_media={len(output_media)}"]
            if empty:
                details.append("empty=" + ",".join(empty))
            raise ValueError("master media integrity failed: " + "; ".join(details))
    except zipfile.BadZipFile as e:
        raise ValueError(f"invalid xlsx while checking master media: {e}")


@router.get("/record/{record_id}/excel")
def export_record(
    record_id:int,
    db:Session=Depends(get_db),
    u=Depends(get_current_user),
    person_key:str=Depends(require_person_key),
):
    x=db.get(SourceFormRecord,record_id)
    if not x or x.created_by != u.id or x.owner_person_key != person_key:
        raise HTTPException(404,"Record not found")

    try:
        d=json.loads(x.payload_json or "{}")
        d=_normalize_formula_material_codes(d)
    except Exception:
        d={}

    templates={
      "F-RD-001":("F-RD-001_TEMPLATE.xlsx","RD"),
      "F-RD-002":("F-RD-002.xlsx","สูตร"),
      "F-RD-002.1":("F-RD-002.1.xlsx","คิดต้นทุน สรุป"),
      "F-RD-003":("F-RD-003.xlsx","Sheet1"),
      "F-RD-004":("F-RD-004.xlsx","Sheet1"),
      "ADMIN-QP":("ADMIN-QP_MASTER.xlsx","ใบราคา"),
      "ADMIN-INVOICE":("ADMIN-QP_MASTER.xlsx","ใบราคา"),
    }

    if x.form_code not in templates:
        raise HTTPException(400,"Unsupported form")

    fn,sheet=templates[x.form_code]
    src=ROOT/fn
    if not src.exists():
        raise HTTPException(500,f"Template not found: {fn}")

    try:
        if x.form_code in ("ADMIN-QP", "ADMIN-INVOICE"):
            # Do NOT save the visual quotation master through openpyxl.
            # Patch only worksheet cell XML inside a copy of the original XLSX
            # package so logos/images/drawings remain untouched on Render too.
            output_bytes = _admin_qp_export_preserve_master(src, d, is_invoice=(x.form_code == "ADMIN-INVOICE"))
            _verify_master_media_preserved(src, output_bytes)
            output = BytesIO(output_bytes)
            output.seek(0)
        else:
            wb=load_workbook(src)
            if sheet not in wb.sheetnames:
                raise HTTPException(500,f"Sheet not found: {sheet}")
            ws=wb[sheet]

            if x.form_code=="F-RD-001":
                fill_001(ws,d)
            elif x.form_code=="F-RD-002":
                fill_formula(ws,d,False)
            elif x.form_code=="F-RD-002.1":
                fill_formula(ws,d,True)
            elif x.form_code=="F-RD-003":
                fill_003(ws,d)
            elif x.form_code=="F-RD-004":
                fill_004(ws,d)
            else:
                raise HTTPException(400,"Unsupported form")

            fill_manual_cells(ws,d,x.form_code=="F-RD-002.1")
            if x.form_code=="F-RD-002.1":
                fill_tester_qty_formula(ws,d)
                fill_margin_percent_formula(ws,d)
            output=BytesIO()
            wb.save(output)
            output.seek(0)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500,f"Excel export failed after merged-cell handling: {type(e).__name__}: {e}")

    safe_record=re.sub(r'[^A-Za-z0-9._-]+','_',str(x.record_no or x.id)).strip('_') or str(x.id)
    filename=f"{x.form_code}_{safe_record}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )



@router.get("/diagnostics/source-form-schema")
def source_form_schema_diagnostics(
    db: Session = Depends(get_db),
    u=Depends(get_current_user),
):
    from sqlalchemy import text
    try:
        dialect=db.bind.dialect.name
        result={"dialect":dialect}
        if dialect=="postgresql":
            cols=db.execute(text("""
                SELECT column_name, data_type, is_nullable
                FROM information_schema.columns
                WHERE table_name='source_form_records'
                ORDER BY ordinal_position
            """)).mappings().all()
            fks=db.execute(text("""
                SELECT
                  kcu.column_name,
                  ccu.table_name AS foreign_table_name,
                  ccu.column_name AS foreign_column_name
                FROM information_schema.table_constraints AS tc
                JOIN information_schema.key_column_usage AS kcu
                  ON tc.constraint_name = kcu.constraint_name
                JOIN information_schema.constraint_column_usage AS ccu
                  ON ccu.constraint_name = tc.constraint_name
                WHERE tc.constraint_type='FOREIGN KEY'
                  AND tc.table_name='source_form_records'
            """)).mappings().all()
            result["columns"]=[dict(x) for x in cols]
            result["foreign_keys"]=[dict(x) for x in fks]
        else:
            cols=db.execute(text("PRAGMA table_info(source_form_records)")).fetchall()
            result["columns"]=[list(x) for x in cols]
        return result
    except Exception as e:
        raise HTTPException(500,f"Diagnostic failed: {type(e).__name__}: {e}")

@router.get("/aliases/{supplement_code}")
def aliases(
    supplement_code: str,
    db: Session = Depends(get_db),
    u=Depends(get_current_user),
    
):
    rows = db.scalars(
        select(SupplementAlias)
        .where(SupplementAlias.supplement_code == supplement_code)
        .order_by(SupplementAlias.id)
    ).all()
    return [{
        "primary_name": x.primary_name,
        "alternate_name": x.alternate_name
    } for x in rows]
