import json
import re
from datetime import date
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.session import get_db
from app.core.security import get_current_user, require_roles
from app.core.thai_baht import thai_baht_text
from app.models.entities import PurchaseDocument

router = APIRouter(prefix="/api/purchase-docs", tags=["Purchase Documents"])

DOC_TYPES = {"PO", "PR"}


def _check_doc_type(doc_type: str) -> str:
    d = (doc_type or "").strip().upper()
    if d not in DOC_TYPES:
        raise HTTPException(404, "Unsupported document type (expected PO or PR)")
    return d


def _validate_doc_data(doc_type: str, data: dict):
    """Server-side mirror of the frontend's save-time validation -- the
    frontend check is for instant feedback, this is what actually protects
    the data (a request that skips the browser, or a future non-browser
    client, must not be able to save an empty document)."""
    items = data.get("items") or []
    if doc_type == "PO":
        if not str(data.get("supplier_name") or "").strip() and not str(data.get("supplier_code") or "").strip():
            raise HTTPException(400, "กรุณาใส่ผู้จำหน่าย (ชื่อ หรือ รหัสผู้จำหน่าย) ก่อนบันทึก")
        if not any(str(x.get("description") or "").strip() for x in items):
            raise HTTPException(400, "กรุณาใส่รายการสินค้าอย่างน้อย 1 รายการ ก่อนบันทึก")
    else:
        if not any(str(x.get("material_code") or "").strip() or str(x.get("description") or "").strip() for x in items):
            raise HTTPException(400, "กรุณาใส่รายการวัตถุดิบอย่างน้อย 1 รายการ (รหัสสินค้าหรือรายละเอียด) ก่อนบันทึก")


class DocSave(BaseModel):
    doc_no: str
    status: str = "DRAFT"
    data: dict
    linked_reference: str | None = None


def _serialize(x: PurchaseDocument) -> dict:
    return {
        "id": x.id,
        "doc_type": x.doc_type,
        "doc_no": x.doc_no,
        "status": x.status,
        "data": json.loads(x.payload_json or "{}"),
        "linked_reference": x.linked_reference,
        "created_by_name": x.created_by_name,
        "created_at": x.created_at,
        "updated_at": x.updated_at,
    }


@router.post("/{doc_type}")
def save_doc(
    doc_type: str,
    p: DocSave,
    db: Session = Depends(get_db),
    u=Depends(get_current_user),
):
    d = _check_doc_type(doc_type)
    if not p.data.get("date"):
        p.data["date"] = date.today().isoformat()
    _validate_doc_data(d, p.data)

    x = PurchaseDocument(
        doc_type=d,
        doc_no=p.doc_no,
        status=p.status,
        payload_json=json.dumps(p.data, ensure_ascii=False, default=str),
        created_by=u.id,
        created_by_name=u.full_name,
        linked_reference=(p.linked_reference or "").strip() or None,
    )
    db.add(x)
    db.commit()
    db.refresh(x)
    return _serialize(x)


@router.get("/{doc_type}")
def list_docs(
    doc_type: str,
    db: Session = Depends(get_db),
    u=Depends(get_current_user),
):
    """Department-shared listing -- same visibility model as Customers/Suppliers:
    any authenticated user can see every PO/PR, not just their own."""
    d = _check_doc_type(doc_type)
    rows = db.scalars(
        select(PurchaseDocument)
        .where(PurchaseDocument.doc_type == d)
        .order_by(PurchaseDocument.id.desc())
    ).all()
    return [_serialize(x) for x in rows]


@router.get("/record/{record_id}")
def get_doc(record_id: int, db: Session = Depends(get_db), u=Depends(get_current_user)):
    x = db.get(PurchaseDocument, record_id)
    if not x:
        raise HTTPException(404, "Record not found")
    return _serialize(x)


@router.put("/record/{record_id}")
def update_doc(
    record_id: int,
    p: DocSave,
    db: Session = Depends(get_db),
    u=Depends(get_current_user),
):
    x = db.get(PurchaseDocument, record_id)
    if not x:
        raise HTTPException(404, "Record not found")
    if not p.data.get("date"):
        p.data["date"] = date.today().isoformat()
    _validate_doc_data(x.doc_type, p.data)
    x.doc_no = p.doc_no
    x.status = p.status
    x.payload_json = json.dumps(p.data, ensure_ascii=False, default=str)
    if p.linked_reference is not None:
        x.linked_reference = p.linked_reference.strip() or None
    db.commit()
    db.refresh(x)
    return _serialize(x)


@router.delete("/record/{record_id}")
def delete_doc(
    record_id: int,
    db: Session = Depends(get_db),
    u=Depends(require_roles("ADMIN", "PURCHASE", "STOCK")),
):
    x = db.get(PurchaseDocument, record_id)
    if not x:
        raise HTTPException(404, "Record not found")
    db.delete(x)
    db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# Excel export
#
# Unlike every other form in the app, PO/PR have no real Excel master to
# patch cell-by-cell (they were built from screenshots of paper documents,
# not sourced from an .xlsx file -- see PR that introduced them). So this
# builds a fresh, readable workbook that mirrors the on-screen layout,
# rather than reusing the "preserve master, patch values" pattern used for
# the other exact forms.
# ---------------------------------------------------------------------------

_THIN = Side(style="thin", color="999999")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEAD_FONT = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=14)
_WRAP = Alignment(wrap_text=True, vertical="top")


def _label_value(ws, row, label_col, label, value, value_col=None):
    lc = get_column_letter(label_col)
    vc = get_column_letter(value_col or label_col + 1)
    ws[f"{lc}{row}"] = label
    ws[f"{lc}{row}"].font = _HEAD_FONT
    ws[f"{vc}{row}"] = value or ""
    ws[f"{vc}{row}"].border = _BORDER


def _build_po_workbook(doc_no: str, data: dict) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "PO"
    for col, width in zip("ABCDEF", [4, 34, 10, 10, 14, 14]):
        ws.column_dimensions[col].width = width

    ws.merge_cells("A1:F1")
    ws["A1"] = "ใบสั่งซื้อ (Purchase Order)"
    ws["A1"].font = _TITLE_FONT

    r = 3
    _label_value(ws, r, 1, "เลขที่", doc_no); r += 1
    _label_value(ws, r, 1, "วันที่", data.get("date")); r += 1
    _label_value(ws, r, 1, "ครบกำหนด", data.get("due_date")); r += 1
    _label_value(ws, r, 1, "ผู้สั่งซื้อ", data.get("buyer_name")); r += 1
    _label_value(ws, r, 1, "อ้างอิง (เลขที่ PR)", data.get("reference")); r += 1
    _label_value(ws, r, 1, "ผู้ติดต่อ", data.get("contact_person")); r += 1
    _label_value(ws, r, 1, "เบอร์โทร", data.get("contact_phone")); r += 1
    _label_value(ws, r, 1, "รหัสผู้จำหน่าย", data.get("supplier_code")); r += 1
    _label_value(ws, r, 1, "ผู้จำหน่าย", data.get("supplier_name")); r += 1
    _label_value(ws, r, 1, "ที่อยู่ผู้จำหน่าย", data.get("supplier_address")); r += 1
    _label_value(ws, r, 1, "เลขประจำตัวผู้เสียภาษี", data.get("supplier_tax_id")); r += 1

    r += 1
    headers = ["#", "รายละเอียด", "จำนวน", "หน่วย", "ราคาต่อหน่วย", "ยอดรวม"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = _HEAD_FONT
        c.border = _BORDER
        c.alignment = Alignment(horizontal="center")
    table_head_row = r
    r += 1

    items = data.get("items") or []
    subtotal = 0.0
    for i, item in enumerate(items, start=1):
        if not str(item.get("description") or "").strip():
            continue
        qty = float(item.get("quantity") or 0) if str(item.get("quantity") or "").strip() else 0
        price = float(item.get("unit_price") or 0) if str(item.get("unit_price") or "").strip() else 0
        amount = qty * price
        subtotal += amount
        vals = [i, item.get("description") or "", item.get("quantity") or "", item.get("unit") or "",
                item.get("unit_price") or "", round(amount, 2) if amount else ""]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.border = _BORDER
            if col == 2:
                c.alignment = _WRAP
        r += 1
    if r == table_head_row + 1:
        # no rows written -- keep at least one bordered blank row for readability
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = _BORDER
        r += 1

    vat = subtotal * 0.07
    grand_total = subtotal + vat
    r += 1
    _label_value(ws, r, 4, "รวมเป็นเงิน", round(subtotal, 2), value_col=5); r += 1
    _label_value(ws, r, 4, "ภาษีมูลค่าเพิ่ม 7%", round(vat, 2), value_col=5); r += 1
    _label_value(ws, r, 4, "จำนวนเงินรวมทั้งสิ้น", round(grand_total, 2), value_col=5); r += 1
    ws.merge_cells(f"A{r}:F{r}")
    ws[f"A{r}"] = f"({thai_baht_text(grand_total)})"
    r += 2

    _label_value(ws, r, 1, "ผู้ซื้อ", data.get("buyer_sign")); r += 1
    _label_value(ws, r, 1, "วันที่ (ผู้ซื้อ)", data.get("buyer_sign_date")); r += 1
    _label_value(ws, r, 1, "ผู้อนุมัติ", data.get("approver_sign")); r += 1
    _label_value(ws, r, 1, "วันที่ (ผู้อนุมัติ)", data.get("approver_sign_date")); r += 1

    return wb


def _build_pr_workbook(doc_no: str, data: dict) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "PR"
    widths = [6, 12, 24, 8, 8, 14, 18, 12, 14, 12]
    for col, width in zip("ABCDEFGHIJ", widths):
        ws.column_dimensions[col].width = width

    ws.merge_cells("A1:J1")
    ws["A1"] = "ใบขอซื้อ (Purchase Request)"
    ws["A1"].font = _TITLE_FONT

    r = 3
    _label_value(ws, r, 1, "เลขที่แบบฟอร์ม", data.get("form_no")); r += 1
    _label_value(ws, r, 1, "แก้ไขครั้งที่", data.get("revision_no")); r += 1
    _label_value(ws, r, 1, "เลขที่ PR", doc_no); r += 1
    _label_value(ws, r, 1, "วันที่", data.get("date")); r += 1
    _label_value(ws, r, 1, "เวลา", data.get("time")); r += 1
    _label_value(ws, r, 1, "เตรียมโดย", data.get("prepared_by")); r += 1
    _label_value(ws, r, 1, "อนุมัติโดย", data.get("approved_by")); r += 1
    _label_value(ws, r, 1, "อ้างอิงสูตร/ผลิตภัณฑ์", data.get("product_ref")); r += 1

    r += 1
    headers = ["ลำดับ", "รหัสสินค้า", "รายละเอียด", "จำนวน", "หน่วย", "เลขที่ใบสั่งผลิต",
               "ชื่อผลิตภัณฑ์/แผนก", "เลขที่ PO", "หมายเหตุ", "วันที่รับเข้า"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = _HEAD_FONT
        c.border = _BORDER
        c.alignment = Alignment(horizontal="center")
    table_head_row = r
    r += 1

    items = data.get("items") or []
    subs = ["material_code", "description", "quantity", "unit", "production_order_no",
            "product_name", "po_no", "note", "received_date"]
    for i, item in enumerate(items, start=1):
        if not any(str(item.get(k) or "").strip() for k in subs):
            continue
        vals = [i] + [item.get(k) or "" for k in subs]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.border = _BORDER
            if col == 3:
                c.alignment = _WRAP
        r += 1
    if r == table_head_row + 1:
        for col in range(1, 11):
            ws.cell(row=r, column=col).border = _BORDER
        r += 1

    r += 1
    for key, label in [
        ("requester", "ผู้ขอซื้อ"),
        ("warehouse_officer", "จนท.คลังสินค้า"),
        ("purchasing_officer", "เจ้าหน้าที่จัดซื้อ"),
        ("reviewer", "ผู้ตรวจสอบ (ผจก.แผนก)"),
        ("warehouse_manager", "ผจก.คลังสินค้า"),
    ]:
        name = data.get(f"sign_{key}") or ""
        sign_date = data.get(f"sign_{key}_date") or ""
        _label_value(ws, r, 1, label, name, value_col=3)
        _label_value(ws, r, 6, "วันที่", sign_date, value_col=7)
        r += 1

    return wb


@router.get("/record/{record_id}/excel")
def export_doc_excel(
    record_id: int,
    db: Session = Depends(get_db),
    u=Depends(get_current_user),
):
    x = db.get(PurchaseDocument, record_id)
    if not x:
        raise HTTPException(404, "Record not found")
    data = json.loads(x.payload_json or "{}")

    try:
        wb = _build_po_workbook(x.doc_no, data) if x.doc_type == "PO" else _build_pr_workbook(x.doc_no, data)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
    except Exception as e:
        raise HTTPException(500, f"Excel export failed: {type(e).__name__}: {e}")

    safe_doc_no = re.sub(r"[^A-Za-z0-9._-]+", "_", str(x.doc_no or x.id)).strip("_") or str(x.id)
    filename = f"{x.doc_type}_{safe_doc_no}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
